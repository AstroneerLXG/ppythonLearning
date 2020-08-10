from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
from PIL import Image  # 图像处理库PythonImageLibrary

newWorkbook = Workbook()  # 新建Excel工作簿
newSheet = newWorkbook.active  # 打开自带的sheet
im = Image.open(r'C:\Users\aby\OneDrive\图片\表情包\封面图.png')
imWidth = im.size[0]
imHeight = im.size[1]
allPixelData = im.load()  # 导出像素

for row in range(1, imHeight):
    for col in range(1, imWidth):
        cell = newSheet.cell(column=col, row=row)  # 获取单元格
        pixelData = allPixelData[col - 1, row - 1]  # 获取RGB色彩数据
        pixelColor = 'FF%02X%02X%02X' % (pixelData[0], pixelData[1], pixelData[2])  # RGB三通道数值用字符串插入的方式转换为HEX格式
        fill = PatternFill(patternType='solid', fgColor=Color(rgb=pixelColor))  # 设置填充样式和色彩
        cell.fill = fill  # 填充操作
    newSheet.row_dimensions[row].height = 6  # 设置单元格的高

for col in range(1, imWidth):
    newSheet.column_dimensions[get_column_letter(col)].width = 1  # 设置单元格的宽

newWorkbook.save('C:/Users/aby/Desktop/QRpicture.xlsx')
